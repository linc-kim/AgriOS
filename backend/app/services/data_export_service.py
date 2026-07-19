"""
Greena — Dataset Export Service (Module 11).

Complements export_service, which produces whole-farm PDF/Excel/CSV bundles.
This module exports a *single dataset* in a chosen format, and — critically —
in a shape that import_service can read back.

Round-tripping is the point: the columns emitted here are the columns the
importer expects, so export → edit in Excel → import is a supported workflow
rather than a coincidence.
"""

import csv
import io
import json
import logging
import time
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ValidationException
from app.models.auth import User
from app.models.farm import Farm
from app.models.finance import Expense, ExpenseCategory, RevenueRecord
from app.models.flock import DailyLog, Flock
from app.models.inventory import InventoryItem
from app.models.production import ExportJob

logger = logging.getLogger(__name__)

SUPPORTED_FORMATS = ("csv", "excel", "json")


def _fmt(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    return value


# ── Dataset builders ──────────────────────────────────────────────────────────
# Each returns (column order, rows). Column names mirror the import templates so
# an exported file can be edited and imported straight back.

async def _daily_logs(db: AsyncSession, farm: Farm) -> tuple[list[str], list[dict]]:
    result = await db.execute(
        select(DailyLog, Flock.name)
        .join(Flock, Flock.id == DailyLog.flock_id, isouter=True)
        .where(DailyLog.farm_id == farm.id, DailyLog.deleted_at.is_(None))
        .order_by(DailyLog.log_date.desc())
    )
    columns = ["log_date", "flock_name", "mortality_count", "culls",
               "feed_consumed_kg", "water_litres", "morning_count", "notes"]
    rows = [
        {
            "log_date": _fmt(log.log_date),
            "flock_name": flock_name or "",
            "mortality_count": log.mortality_count,
            "culls": log.culls,
            "feed_consumed_kg": _fmt(log.feed_consumed_kg),
            "water_litres": _fmt(log.water_litres),
            "morning_count": log.morning_count,
            "notes": log.notes or "",
        }
        for log, flock_name in result
    ]
    return columns, rows


async def _expenses(db: AsyncSession, farm: Farm) -> tuple[list[str], list[dict]]:
    result = await db.execute(
        select(Expense, ExpenseCategory.name)
        .join(ExpenseCategory, ExpenseCategory.id == Expense.category_id, isouter=True)
        .where(Expense.farm_id == farm.id, Expense.deleted_at.is_(None))
        .order_by(Expense.expense_date.desc())
    )
    columns = ["expense_date", "category", "amount", "description",
               "payment_method", "supplier", "notes"]
    rows = [
        {
            "expense_date": _fmt(e.expense_date),
            "category": category or "other",
            "amount": _fmt(e.amount),
            "description": e.description or "",
            "payment_method": e.payment_method or "",
            "supplier": e.supplier or "",
            "notes": e.notes or "",
        }
        for e, category in result
    ]
    return columns, rows


async def _revenue(db: AsyncSession, farm: Farm) -> tuple[list[str], list[dict]]:
    result = await db.execute(
        select(RevenueRecord, Flock.name)
        .join(Flock, Flock.id == RevenueRecord.flock_id, isouter=True)
        .where(RevenueRecord.farm_id == farm.id, RevenueRecord.deleted_at.is_(None))
        .order_by(RevenueRecord.revenue_date.desc())
    )
    columns = ["revenue_date", "flock_name", "revenue_type", "amount",
               "quantity", "unit_price", "buyer_name", "notes"]
    rows = [
        {
            "revenue_date": _fmt(r.revenue_date),
            "flock_name": flock_name or "",
            "revenue_type": r.revenue_type,
            "amount": _fmt(r.amount),
            "quantity": _fmt(r.quantity),
            "unit_price": _fmt(r.unit_price),
            "buyer_name": r.buyer_name or "",
            "notes": r.notes or "",
        }
        for r, flock_name in result
    ]
    return columns, rows


async def _inventory_items(db: AsyncSession, farm: Farm) -> tuple[list[str], list[dict]]:
    result = await db.execute(
        select(InventoryItem)
        .where(InventoryItem.farm_id == farm.id, InventoryItem.deleted_at.is_(None))
        .order_by(InventoryItem.name)
    )
    columns = ["name", "category", "unit", "quantity", "reorder_level",
               "purchase_price", "description"]
    rows = [
        {
            "name": item.name,
            "category": item.category,
            "unit": item.unit,
            "quantity": _fmt(item.quantity),
            "reorder_level": _fmt(item.reorder_level),
            "purchase_price": _fmt(item.purchase_price),
            "description": item.description or "",
        }
        for item in result.scalars().all()
    ]
    return columns, rows


async def _flocks(db: AsyncSession, farm: Farm) -> tuple[list[str], list[dict]]:
    result = await db.execute(
        select(Flock)
        .where(Flock.farm_id == farm.id, Flock.deleted_at.is_(None))
        .order_by(Flock.placement_date.desc())
    )
    columns = ["name", "species_key", "breed", "initial_count",
               "placement_date", "status"]
    rows = [
        {
            "name": f.name,
            "species_key": f.species_key,
            "breed": f.breed or "",
            "initial_count": f.initial_count,
            "placement_date": _fmt(f.placement_date),
            "status": f.status,
        }
        for f in result.scalars().all()
    ]
    return columns, rows


DATASETS = {
    "daily_logs": _daily_logs,
    "expenses": _expenses,
    "revenue": _revenue,
    "inventory_items": _inventory_items,
    "flocks": _flocks,
}


# ── Renderers ─────────────────────────────────────────────────────────────────

def _render_csv(columns: list[str], rows: list[dict]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    # utf-8-sig: Excel misreads plain UTF-8 CSVs, mangling any non-ASCII name.
    return buffer.getvalue().encode("utf-8-sig")


def _render_json(dataset: str, farm: Farm, columns: list[str], rows: list[dict]) -> bytes:
    payload = {
        "dataset": dataset,
        "farm": {"id": str(farm.id), "name": farm.name},
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "columns": columns,
        "count": len(rows),
        # "data" is the key the importer looks for, so this file imports back in.
        "data": rows,
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _render_excel(dataset: str, columns: list[str], rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = dataset[:31]  # Excel caps sheet names at 31 characters

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="076524")  # Greena green

    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append([row.get(c, "") for c in columns])

    for i, column in enumerate(columns, start=1):
        longest = max(
            [len(str(column))] + [len(str(r.get(column, ""))) for r in rows[:200]]
        )
        sheet.column_dimensions[get_column_letter(i)].width = min(longest + 2, 40)

    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── Entry point ───────────────────────────────────────────────────────────────

async def export_dataset(
    db: AsyncSession,
    farm: Farm,
    dataset: str,
    export_format: str,
    user: User | None = None,
) -> tuple[bytes, str, str]:
    """
    Export one dataset. Returns (content, filename, media_type).

    The ExportJob row is written whatever the outcome, so the Exports page shows
    failures rather than silently missing entries.
    """
    started = time.perf_counter()

    builder = DATASETS.get(dataset)
    if not builder:
        raise ValidationException(
            f"Unknown dataset '{dataset}'. Available: {', '.join(sorted(DATASETS))}."
        )
    if export_format not in SUPPORTED_FORMATS:
        raise ValidationException(
            f"Unsupported format '{export_format}'. Use: {', '.join(SUPPORTED_FORMATS)}."
        )

    job = ExportJob(
        farm_id=farm.id,
        dataset=dataset,
        export_format=export_format,
        status="running",
        created_by=user.id if user else None,
    )
    db.add(job)
    await db.flush()

    try:
        columns, rows = await builder(db, farm)

        if export_format == "csv":
            content = _render_csv(columns, rows)
            media_type = "text/csv"
            extension = "csv"
        elif export_format == "json":
            content = _render_json(dataset, farm, columns, rows)
            media_type = "application/json"
            extension = "json"
        else:
            content = _render_excel(dataset, columns, rows)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"

        job.row_count = len(rows)
        job.size_bytes = len(content)
        job.status = "success"
        job.duration_ms = int((time.perf_counter() - started) * 1000)
        await db.commit()

        from app.services.metrics_service import registry

        registry.record_event("export_created")

        safe_farm = "".join(c if c.isalnum() or c in "-_" else "_" for c in farm.name)[:40]
        filename = f"{safe_farm}_{dataset}_{datetime.now(timezone.utc):%Y%m%d_%H%M}.{extension}"
        return content, filename, media_type

    except Exception as exc:
        job.status = "failed"
        job.error = str(exc)[:2000]
        job.duration_ms = int((time.perf_counter() - started) * 1000)
        await db.commit()

        from app.services.metrics_service import registry

        registry.record_event("export_failed")
        logger.error("Export failed for farm %s dataset %s: %s", farm.id, dataset, exc)
        raise


async def list_export_jobs(db: AsyncSession, farm_id: uuid.UUID) -> list[ExportJob]:
    result = await db.execute(
        select(ExportJob)
        .where(ExportJob.farm_id == farm_id, ExportJob.deleted_at.is_(None))
        .order_by(ExportJob.created_at.desc())
        .limit(50)
    )
    return list(result.scalars().all())
