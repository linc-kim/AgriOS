"""
Greena — Import Service (Module 11).

Bulk-loads farm data from CSV, Excel or JSON.

The design principle is that a farmer should never be able to corrupt a season
of records with one bad upload. So:

  * Every import validates the whole file before writing anything.
  * Dry run is the default; committing is an explicit second step.
  * Validation failures are reported per row and per field, with the row number
    from the original file, so a spreadsheet can actually be corrected.
  * A file with any invalid row still imports the valid ones only if the caller
    asks for it; otherwise nothing is written.
"""

import csv
import io
import json
import logging
import time
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ValidationException
from app.models.auth import User
from app.models.farm import Farm
from app.models.finance import Expense, RevenueRecord
from app.models.flock import DailyLog, Flock
from app.models.inventory import InventoryItem
from app.models.production import ImportJob

logger = logging.getLogger(__name__)

MAX_ROWS = 10_000
MAX_STORED_ERRORS = 200


# ── Field coercion ────────────────────────────────────────────────────────────

def _to_date(value: Any, field: str) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        raise ValueError(f"{field} is required")
    # Excel and hand-typed sheets vary; accept the common Kenyan orderings.
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        raise ValueError(f"{field} '{text}' is not a recognised date (use YYYY-MM-DD)")


def _to_decimal(value: Any, field: str, required: bool = True) -> Decimal | None:
    text = str(value).strip() if value is not None else ""
    if not text:
        if required:
            raise ValueError(f"{field} is required")
        return None
    # Tolerate thousands separators and currency prefixes from spreadsheets.
    cleaned = text.replace(",", "").replace("KES", "").replace("Ksh", "").strip()
    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field} '{text}' is not a number")
    if parsed < 0:
        raise ValueError(f"{field} cannot be negative")
    return parsed


def _to_int(value: Any, field: str, required: bool = True) -> int | None:
    text = str(value).strip() if value is not None else ""
    if not text:
        if required:
            raise ValueError(f"{field} is required")
        return None
    try:
        parsed = int(float(text.replace(",", "")))
    except (ValueError, TypeError):
        raise ValueError(f"{field} '{text}' is not a whole number")
    if parsed < 0:
        raise ValueError(f"{field} cannot be negative")
    return parsed


def _to_str(value: Any, field: str, required: bool = False, max_length: int = 200) -> str | None:
    text = str(value).strip() if value is not None else ""
    if not text:
        if required:
            raise ValueError(f"{field} is required")
        return None
    if len(text) > max_length:
        raise ValueError(f"{field} exceeds {max_length} characters")
    return text


# ── Entity definitions ────────────────────────────────────────────────────────

class EntitySpec:
    """How one importable entity maps from file columns to model kwargs."""

    def __init__(
        self,
        model: Any,
        columns: list[str],
        required: list[str],
        builder: Callable[[dict, Farm, dict], dict],
        needs_flock: bool = False,
        needs_categories: bool = False,
    ) -> None:
        self.model = model
        self.columns = columns
        self.required = required
        self.builder = builder
        self.needs_flock = needs_flock
        self.needs_categories = needs_categories


def _build_daily_log(row: dict, farm: Farm, ctx: dict) -> dict:
    flock_id = _resolve_flock(row, ctx)
    return {
        "farm_id": farm.id,
        "flock_id": flock_id,
        "log_date": _to_date(row.get("log_date"), "log_date"),
        "mortality_count": _to_int(row.get("mortality_count"), "mortality_count", required=False) or 0,
        "culls": _to_int(row.get("culls"), "culls", required=False) or 0,
        "feed_consumed_kg": _to_decimal(row.get("feed_consumed_kg"), "feed_consumed_kg", required=False) or Decimal("0"),
        "water_litres": _to_decimal(row.get("water_litres"), "water_litres", required=False),
        "morning_count": _to_int(row.get("morning_count"), "morning_count", required=False),
        "notes": _to_str(row.get("notes"), "notes", max_length=1000),
    }


def _build_expense(row: dict, farm: Farm, ctx: dict) -> dict:
    return {
        "farm_id": farm.id,
        # category_id and description are both NOT NULL on the model.
        "category_id": _resolve_category(row, ctx),
        "expense_date": _to_date(row.get("expense_date"), "expense_date"),
        "amount": _to_decimal(row.get("amount"), "amount"),
        "description": _to_str(row.get("description"), "description", required=True, max_length=500),
        "payment_method": _to_str(row.get("payment_method"), "payment_method", max_length=20),
        "supplier": _to_str(row.get("supplier"), "supplier", max_length=200),
        "notes": _to_str(row.get("notes"), "notes", max_length=1000),
    }


def _resolve_category(row: dict, ctx: dict) -> uuid.UUID:
    """
    Map a category name or slug from the file to an expense category.

    Matched case-insensitively against both the display name ("Feed Purchase")
    and the slug ("feed_purchase"), since a farmer's sheet will have whichever
    they saw in the UI. An unrecognised value falls back to "other" rather than
    failing the row — a mis-categorised expense is recoverable, a rejected
    import of a season's records is a much worse outcome.
    """
    raw = str(row.get("category") or row.get("category_name") or "").strip().lower()
    if raw:
        match = ctx["categories"].get(raw)
        if match:
            return match
    fallback = ctx["categories"].get("other")
    if fallback is None:
        raise ValueError("no expense categories are configured for this farm")
    return fallback


def _build_revenue(row: dict, farm: Farm, ctx: dict) -> dict:
    return {
        "farm_id": farm.id,
        # Revenue is always attributed to a flock (flock_id is NOT NULL), so the
        # file must identify one — same resolution rules as daily logs.
        "flock_id": _resolve_flock(row, ctx),
        "revenue_date": _to_date(row.get("revenue_date"), "revenue_date"),
        "revenue_type": _to_str(row.get("revenue_type"), "revenue_type", required=True, max_length=30) or "other",
        "amount": _to_decimal(row.get("amount"), "amount"),
        "quantity": _to_decimal(row.get("quantity"), "quantity", required=False),
        "unit_price": _to_decimal(row.get("unit_price"), "unit_price", required=False),
        "buyer_name": _to_str(row.get("buyer_name"), "buyer_name", max_length=200),
        "notes": _to_str(row.get("notes"), "notes", max_length=1000),
    }


def _build_inventory_item(row: dict, farm: Farm, ctx: dict) -> dict:
    return {
        "farm_id": farm.id,
        "name": _to_str(row.get("name"), "name", required=True),
        "category": _to_str(row.get("category"), "category", max_length=50) or "other",
        "unit": _to_str(row.get("unit"), "unit", max_length=20) or "unit",
        # The column is `quantity`, and cost is `purchase_price` — not the
        # quantity_on_hand / unit_cost names a spreadsheet might use, so the
        # file's friendlier headers are mapped here.
        "quantity": _to_decimal(row.get("quantity") or row.get("quantity_on_hand"),
                                "quantity", required=False) or Decimal("0"),
        "reorder_level": _to_decimal(row.get("reorder_level"), "reorder_level", required=False),
        "purchase_price": _to_decimal(row.get("purchase_price") or row.get("unit_cost"),
                                      "purchase_price", required=False),
        "description": _to_str(row.get("description"), "description", max_length=500),
    }


def _resolve_flock(row: dict, ctx: dict) -> uuid.UUID:
    """
    Resolve the flock a row belongs to, by id or by name.

    Names are what appear in a farmer's spreadsheet; ids are what an export
    round-trip produces. Both are accepted, and an unknown value fails the row
    rather than silently attaching data to the wrong flock.
    """
    raw = str(row.get("flock_id") or row.get("flock") or row.get("flock_name") or "").strip()
    if not raw:
        raise ValueError("flock_id or flock_name is required")

    try:
        parsed = uuid.UUID(raw)
        if parsed in ctx["flock_ids"]:
            return parsed
        raise ValueError(f"flock '{raw}' does not belong to this farm")
    except ValueError as exc:
        if "does not belong" in str(exc):
            raise

    match = ctx["flocks_by_name"].get(raw.lower())
    if match is None:
        known = ", ".join(sorted(ctx["flocks_by_name"].keys())[:5]) or "none"
        raise ValueError(f"no flock named '{raw}' on this farm (known: {known})")
    return match


ENTITY_SPECS: dict[str, EntitySpec] = {
    "daily_logs": EntitySpec(
        DailyLog,
        columns=["log_date", "flock_name", "mortality_count", "culls",
                 "feed_consumed_kg", "water_litres", "morning_count", "notes"],
        required=["log_date", "flock_name"],
        builder=_build_daily_log,
        needs_flock=True,
    ),
    "expenses": EntitySpec(
        Expense,
        columns=["expense_date", "category", "amount", "description",
                 "payment_method", "supplier", "notes"],
        required=["expense_date", "amount", "description"],
        builder=_build_expense,
        needs_categories=True,
    ),
    "revenue": EntitySpec(
        RevenueRecord,
        columns=["revenue_date", "flock_name", "revenue_type", "amount", "quantity",
                 "unit_price", "buyer_name", "notes"],
        required=["revenue_date", "flock_name", "revenue_type", "amount"],
        builder=_build_revenue,
        needs_flock=True,
    ),
    "inventory_items": EntitySpec(
        InventoryItem,
        columns=["name", "category", "unit", "quantity_on_hand", "reorder_level", "unit_cost"],
        required=["name"],
        builder=_build_inventory_item,
    ),
}


def entity_template(entity: str) -> str:
    """A header-only CSV a user can fill in — the contract for that entity."""
    spec = ENTITY_SPECS.get(entity)
    if not spec:
        raise ValidationException(f"Unknown import entity '{entity}'.")
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(spec.columns)
    writer.writerow([f"<{c}{'*' if c in spec.required else ''}>" for c in spec.columns])
    return buffer.getvalue()


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_file(content: bytes, source_format: str) -> list[dict]:
    """Parse an uploaded file into a list of row dicts keyed by column name."""
    if source_format == "csv":
        return _parse_csv(content)
    if source_format == "excel":
        return _parse_excel(content)
    if source_format == "json":
        return _parse_json(content)
    raise ValidationException(f"Unsupported format '{source_format}'. Use csv, excel or json.")


def _parse_csv(content: bytes) -> list[dict]:
    try:
        # utf-8-sig strips the BOM Excel writes, which would otherwise become
        # part of the first column's name and break every lookup on it.
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or "").strip(): v for k, v in row.items()} for row in reader]


def _parse_excel(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return []

    parsed = []
    for values in rows:
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # skip blank spacer rows
        parsed.append({header[i]: v for i, v in enumerate(values) if i < len(header)})
    workbook.close()
    return parsed


def _parse_json(content: bytes) -> list[dict]:
    try:
        data = json.loads(content.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValidationException(f"File is not valid JSON: {exc}")

    if isinstance(data, dict):
        # Accept both a bare list and the {"data": [...]} envelope an export emits.
        for key in ("data", "rows", "items", "records"):
            if isinstance(data.get(key), list):
                return data[key]
        raise ValidationException(
            "JSON object must contain a 'data', 'rows', 'items' or 'records' array."
        )
    if not isinstance(data, list):
        raise ValidationException("JSON must be an array of objects.")
    return data


# ── Import ────────────────────────────────────────────────────────────────────

async def run_import(
    db: AsyncSession,
    farm: Farm,
    entity: str,
    content: bytes,
    source_format: str,
    user: User | None = None,
    filename: str | None = None,
    dry_run: bool = True,
    skip_invalid: bool = False,
) -> ImportJob:
    """
    Validate and (optionally) commit an import.

    With skip_invalid=False — the default — a single bad row aborts the whole
    import and nothing is written. That is the safe behaviour for a farmer
    re-uploading a corrected sheet: partial application is what produces
    duplicate and half-missing records.
    """
    started = time.perf_counter()

    spec = ENTITY_SPECS.get(entity)
    if not spec:
        raise ValidationException(
            f"Unknown entity '{entity}'. Supported: {', '.join(sorted(ENTITY_SPECS))}."
        )

    job = ImportJob(
        farm_id=farm.id,
        entity=entity,
        source_format=source_format,
        filename=filename,
        dry_run=dry_run,
        status="running",
        started_at=datetime.now(timezone.utc),
        created_by=user.id if user else None,
    )
    db.add(job)
    await db.flush()

    try:
        rows = parse_file(content, source_format)

        if len(rows) > MAX_ROWS:
            raise ValidationException(
                f"File has {len(rows)} rows, over the {MAX_ROWS} limit. Split it and import in parts."
            )

        ctx = await _build_context(db, farm, spec)

        valid: list[dict] = []
        errors: list[dict] = []

        for index, row in enumerate(rows):
            # +2: one for the header line, one because humans count from 1.
            line = index + 2
            try:
                valid.append(spec.builder(row, farm, ctx))
            except ValueError as exc:
                errors.append({"row": line, "message": str(exc)})
            except Exception as exc:
                errors.append({"row": line, "message": f"Unexpected error: {exc}"})

        job.total_rows = len(rows)
        job.valid_rows = len(valid)
        job.failed_rows = len(errors)
        job.errors = errors[:MAX_STORED_ERRORS]

        if errors and not skip_invalid:
            job.status = "failed"
            job.imported_rows = 0
            job.error = (
                f"{len(errors)} of {len(rows)} rows are invalid. "
                "Fix them, or re-run with skip_invalid to import the valid rows only."
            )
        elif dry_run:
            job.status = "success"
            job.imported_rows = 0
        else:
            for kwargs in valid:
                db.add(spec.model(**kwargs))
            job.imported_rows = len(valid)
            job.status = "success"

        job.completed_at = datetime.now(timezone.utc)
        job.duration_ms = int((time.perf_counter() - started) * 1000)
        await db.commit()
        await db.refresh(job)

        from app.services.metrics_service import registry

        registry.record_event(
            "import_dry_run" if dry_run else
            ("import_failed" if job.status == "failed" else "import_applied")
        )
        logger.info(
            "Import %s: entity=%s rows=%d valid=%d failed=%d dry_run=%s",
            job.id, entity, job.total_rows, job.valid_rows, job.failed_rows, dry_run,
        )

    except Exception as exc:
        await db.rollback()
        # Re-record the job on a clean transaction: the rollback discarded it,
        # and an import that blew up should still leave a trace.
        job = ImportJob(
            farm_id=farm.id,
            entity=entity,
            source_format=source_format,
            filename=filename,
            dry_run=dry_run,
            status="failed",
            error=str(exc)[:2000],
            started_at=datetime.now(timezone.utc),
            completed_at=datetime.now(timezone.utc),
            duration_ms=int((time.perf_counter() - started) * 1000),
            created_by=user.id if user else None,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        from app.services.metrics_service import registry

        registry.record_event("import_failed")
        logger.error("Import failed for farm %s entity %s: %s", farm.id, entity, exc)

    return job


async def _build_context(db: AsyncSession, farm: Farm, spec: EntitySpec) -> dict:
    """
    Pre-load the lookups a row builder needs, once per import rather than per
    row — resolving a flock or category inside the loop would make a 5,000-row
    file 5,000 round trips.
    """
    ctx: dict = {}

    if spec.needs_flock:
        result = await db.execute(
            select(Flock.id, Flock.name).where(
                Flock.farm_id == farm.id, Flock.deleted_at.is_(None)
            )
        )
        rows = list(result)
        ctx["flock_ids"] = {r[0] for r in rows}
        ctx["flocks_by_name"] = {r[1].lower(): r[0] for r in rows}

    if spec.needs_categories:
        from app.models.finance import ExpenseCategory

        result = await db.execute(
            select(ExpenseCategory.id, ExpenseCategory.name, ExpenseCategory.slug).where(
                ExpenseCategory.deleted_at.is_(None),
                (ExpenseCategory.farm_id == farm.id) | (ExpenseCategory.farm_id.is_(None)),
            )
        )
        categories: dict[str, uuid.UUID] = {}
        for cat_id, name, slug in result:
            # Indexed under both spellings so either matches.
            if name:
                categories[name.lower()] = cat_id
            if slug:
                categories[slug.lower()] = cat_id
        ctx["categories"] = categories

    return ctx


async def list_import_jobs(db: AsyncSession, farm_id: uuid.UUID) -> list[ImportJob]:
    result = await db.execute(
        select(ImportJob)
        .where(ImportJob.farm_id == farm_id, ImportJob.deleted_at.is_(None))
        .order_by(ImportJob.created_at.desc())
        .limit(50)
    )
    return list(result.scalars().all())
