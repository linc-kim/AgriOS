"""
Greena — Farm Data Export Service
Generates professional PDF, Excel, and CSV exports of farm data.

Supported formats:
  PDF  — branded report with Greena logo, tables, and summary stats
  XLSX — multi-sheet Excel workbook (one sheet per data type)
  CSV  — flat CSV of daily logs (simple, importable anywhere)

All exports are scoped to a single farm and respect the farm_id
tenancy model — a farmer can only export their own data.
"""

import csv
import io
import uuid
from datetime import datetime
from typing import Any

from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.farm import Farm
from app.models.flock import DailyLog, Flock
from app.models.finance import Expense, FinancialSnapshot, RevenueRecord
from app.models.health import VaccinationRecord


# ── Greena Brand Colours (RGB) ────────────────────────────────────────────────
_GREEN = (22, 163, 74)
_DARK_GREEN = (20, 83, 45)
_LIGHT_GREEN = (240, 253, 244)
_WHITE = (255, 255, 255)
_GRAY = (107, 114, 128)
_BLACK = (17, 24, 39)


# ── Data Loader ───────────────────────────────────────────────────────────────

async def _load_farm_data(
    db: AsyncSession,
    farm_id: uuid.UUID,
) -> dict[str, Any]:
    """Load all farm data needed for any export format."""
    fid = str(farm_id)

    # Farm
    farm_r = await db.execute(select(Farm).where(Farm.id == fid))
    farm = farm_r.scalar_one_or_none()

    # Flocks (all, including closed)
    flock_r = await db.execute(
        select(Flock)
        .where(and_(Flock.farm_id == fid, Flock.deleted_at.is_(None)))
        .order_by(Flock.created_at.asc())
    )
    flocks = flock_r.scalars().all()
    flock_ids = [str(f.id) for f in flocks]

    # Daily logs
    daily_logs: list[Any] = []
    if flock_ids:
        log_r = await db.execute(
            select(DailyLog)
            .where(
                and_(
                    DailyLog.flock_id.in_(flock_ids),
                    DailyLog.deleted_at.is_(None),
                )
            )
            .order_by(DailyLog.log_date.asc())
        )
        daily_logs = log_r.scalars().all()

    # Vaccinations
    vaccinations: list[Any] = []
    if flock_ids:
        vax_r = await db.execute(
            select(VaccinationRecord)
            .where(
                and_(
                    VaccinationRecord.flock_id.in_(flock_ids),
                    VaccinationRecord.deleted_at.is_(None),
                )
            )
            .order_by(VaccinationRecord.administered_date.asc())
        )
        vaccinations = vax_r.scalars().all()

    # Expenses
    exp_r = await db.execute(
        select(Expense)
        .where(and_(Expense.farm_id == fid, Expense.deleted_at.is_(None)))
        .order_by(Expense.expense_date.asc())
    )
    expenses = exp_r.scalars().all()

    # Revenue
    rev_r = await db.execute(
        select(RevenueRecord)
        .where(and_(RevenueRecord.farm_id == fid, RevenueRecord.deleted_at.is_(None)))
        .order_by(RevenueRecord.sale_date.asc())
    )
    revenues = rev_r.scalars().all()

    # Financial snapshots (one per flock)
    snapshots: dict[str, Any] = {}
    if flock_ids:
        snap_r = await db.execute(
            select(FinancialSnapshot).where(
                FinancialSnapshot.flock_id.in_(flock_ids)
            )
        )
        for snap in snap_r.scalars().all():
            snapshots[str(snap.flock_id)] = snap

    return {
        "farm": farm,
        "flocks": flocks,
        "daily_logs": daily_logs,
        "vaccinations": vaccinations,
        "expenses": expenses,
        "revenues": revenues,
        "snapshots": snapshots,
        "exported_at": datetime.utcnow(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _rgb_color(rgb: tuple[int, int, int]):
    from reportlab.lib.colors import Color
    return Color(rgb[0] / 255, rgb[1] / 255, rgb[2] / 255)


async def export_pdf(db: AsyncSession, farm_id: uuid.UUID) -> bytes:
    """Generate a branded PDF report for the farm."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
        PageBreak,
    )
    from reportlab.lib import colors as rl_colors

    data = await _load_farm_data(db, farm_id)
    farm = data["farm"]
    farm_name = farm.name if farm else "Unknown Farm"
    exported_at = data["exported_at"].strftime("%d %B %Y at %H:%M UTC")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm,
        title=f"Greena Report — {farm_name}",
        author="Greena Agricultural Operating System",
    )

    styles = getSampleStyleSheet()
    green = _rgb_color(_GREEN)
    dark_green = _rgb_color(_DARK_GREEN)
    gray = _rgb_color(_GRAY)

    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=dark_green,
                         fontSize=18, spaceAfter=4)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=green,
                         fontSize=13, spaceAfter=4, spaceBefore=10)
    body = ParagraphStyle("Body", parent=styles["Normal"], fontSize=9,
                           spaceAfter=4, leading=14)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8,
                            textColor=gray, leading=11)
    center = ParagraphStyle("Center", parent=styles["Normal"], fontSize=9,
                              alignment=TA_CENTER)

    def tbl_style(header_rows: int = 1) -> TableStyle:
        cmds = [
            ("BACKGROUND",  (0, 0), (-1, header_rows - 1), green),
            ("TEXTCOLOR",   (0, 0), (-1, header_rows - 1), rl_colors.white),
            ("FONTNAME",    (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1),
             [rl_colors.white, _rgb_color(_LIGHT_GREEN)]),
            ("GRID",        (0, 0), (-1, -1), 0.3, _rgb_color(_GRAY)),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ]
        return TableStyle(cmds)

    story = []
    W = A4[0] - 40 * mm  # usable width

    # ── Cover / Header ────────────────────────────────────────────────────────
    story.append(Paragraph("🌿 Greena", ParagraphStyle(
        "Logo", parent=styles["Normal"], fontSize=28, textColor=dark_green,
        spaceAfter=2, alignment=TA_LEFT,
    )))
    story.append(Paragraph("Agricultural Operating System", ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=10, textColor=gray,
        spaceAfter=8,
    )))
    story.append(HRFlowable(width="100%", thickness=2, color=green))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"Farm Report: {farm_name}", h1))
    story.append(Paragraph(f"Generated: {exported_at}", small))
    story.append(Spacer(1, 6 * mm))

    # ── Farm Summary ──────────────────────────────────────────────────────────
    story.append(Paragraph("Farm Overview", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=green))
    story.append(Spacer(1, 2 * mm))

    flocks = data["flocks"]
    active_flocks = [f for f in flocks if f.status == "active"]
    total_birds = sum(getattr(f, "current_bird_count", f.initial_bird_count) or 0
                      for f in active_flocks)

    farm_rows = [
        ["Field", "Value"],
        ["Farm Name", farm_name],
        ["Active Flocks", str(len(active_flocks))],
        ["Total Active Birds", f"{total_birds:,}"],
        ["Total Flocks (all time)", str(len(flocks))],
        ["Report Date", data["exported_at"].strftime("%Y-%m-%d")],
    ]
    t = Table(farm_rows, colWidths=[W * 0.35, W * 0.65])
    t.setStyle(tbl_style(1))
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    # ── Flocks ────────────────────────────────────────────────────────────────
    story.append(Paragraph("Flock Summary", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=green))
    story.append(Spacer(1, 2 * mm))

    flock_map = {str(f.id): f for f in flocks}
    if flocks:
        flock_rows = [["Flock Name", "Status", "Breed", "Initial Birds",
                        "Placement Date", "Days Alive"]]
        for f in flocks:
            days = None
            if f.placement_date:
                days = (datetime.utcnow().date() - f.placement_date).days
            flock_rows.append([
                f.name,
                f.status.upper(),
                getattr(f, "breed", "—") or "—",
                str(f.initial_bird_count),
                str(f.placement_date) if f.placement_date else "—",
                str(days) if days is not None else "—",
            ])
        col_w = W / 6
        t = Table(flock_rows, colWidths=[col_w * 1.5, col_w * 0.8, col_w,
                                          col_w * 0.9, col_w, col_w * 0.8])
        t.setStyle(tbl_style(1))
        story.append(t)
    else:
        story.append(Paragraph("No flocks recorded yet.", body))
    story.append(Spacer(1, 6 * mm))

    # ── Daily Logs ────────────────────────────────────────────────────────────
    story.append(Paragraph("Daily Operations Log", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=green))
    story.append(Spacer(1, 2 * mm))

    daily_logs = data["daily_logs"]
    if daily_logs:
        log_rows = [["Date", "Flock", "Mortality", "Feed (kg)", "Water (L)"]]
        for log in daily_logs[-100:]:  # Last 100 for PDF brevity
            flock = flock_map.get(str(log.flock_id))
            log_rows.append([
                str(log.log_date),
                flock.name if flock else "—",
                str(log.mortality_count),
                f"{float(log.feed_kg):.1f}" if log.feed_kg else "—",
                f"{float(log.water_litres):.1f}"
                    if getattr(log, "water_litres", None) else "—",
            ])
        col_w = W / 5
        t = Table(log_rows, colWidths=[col_w] * 5)
        t.setStyle(tbl_style(1))
        story.append(t)
        if len(daily_logs) > 100:
            story.append(Paragraph(
                f"Showing last 100 of {len(daily_logs)} records. "
                "Download the Excel or CSV export for the full dataset.",
                small,
            ))
    else:
        story.append(Paragraph("No daily logs recorded yet.", body))
    story.append(Spacer(1, 6 * mm))

    # ── Vaccinations ──────────────────────────────────────────────────────────
    story.append(Paragraph("Vaccination Records", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=green))
    story.append(Spacer(1, 2 * mm))

    vaccinations = data["vaccinations"]
    if vaccinations:
        vax_rows = [["Date", "Flock", "Vaccine", "Dose", "Next Due", "Notes"]]
        for v in vaccinations:
            flock = flock_map.get(str(v.flock_id))
            vax_rows.append([
                str(v.administered_date),
                flock.name if flock else "—",
                v.vaccine_name,
                getattr(v, "dose", "—") or "—",
                str(v.next_due_date) if v.next_due_date else "—",
                (getattr(v, "notes", "") or "")[:40],
            ])
        col_w = W / 6
        t = Table(vax_rows, colWidths=[col_w * 0.9, col_w, col_w * 1.1,
                                        col_w * 0.7, col_w * 0.9, col_w * 1.4])
        t.setStyle(tbl_style(1))
        story.append(t)
    else:
        story.append(Paragraph("No vaccination records yet.", body))
    story.append(Spacer(1, 6 * mm))

    # ── Financials ────────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Financial Summary", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=green))
    story.append(Spacer(1, 2 * mm))

    expenses = data["expenses"]
    revenues = data["revenues"]
    total_exp = sum(float(e.amount) for e in expenses)
    total_rev = sum(float(r.amount) for r in revenues)
    net = total_rev - total_exp

    fin_summary = [
        ["Metric", "Amount (KES)"],
        ["Total Revenue", f"KES {total_rev:,.2f}"],
        ["Total Expenses", f"KES {total_exp:,.2f}"],
        ["Net Profit / (Loss)", f"KES {net:,.2f}"],
    ]
    t = Table(fin_summary, colWidths=[W * 0.5, W * 0.5])
    t.setStyle(tbl_style(1))
    story.append(t)
    story.append(Spacer(1, 4 * mm))

    # Expenses table
    if expenses:
        story.append(Paragraph("Expense Records", h2))
        exp_rows = [["Date", "Description / Notes", "Amount (KES)"]]
        for e in expenses:
            exp_rows.append([
                str(e.expense_date),
                (e.notes or "")[:50],
                f"{float(e.amount):,.2f}",
            ])
        t = Table(exp_rows, colWidths=[W * 0.2, W * 0.55, W * 0.25])
        t.setStyle(tbl_style(1))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

    # Revenue table
    if revenues:
        story.append(Paragraph("Revenue Records", h2))
        rev_rows = [["Date", "Type", "Buyer", "Amount (KES)"]]
        for r in revenues:
            rev_rows.append([
                str(r.sale_date),
                r.revenue_type,
                getattr(r, "buyer", "—") or "—",
                f"{float(r.amount):,.2f}",
            ])
        t = Table(rev_rows, colWidths=[W * 0.2, W * 0.25, W * 0.3, W * 0.25])
        t.setStyle(tbl_style(1))
        story.append(t)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 8 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=gray))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "This report was generated by Greena — Agricultural Operating System. "
        "Data is provided for operational reference only. "
        "Always consult a qualified veterinarian for health decisions.",
        small,
    ))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

async def export_excel(db: AsyncSession, farm_id: uuid.UUID) -> bytes:
    """Generate a multi-sheet Excel workbook for the farm."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter

    data = await _load_farm_data(db, farm_id)
    farm = data["farm"]
    farm_name = farm.name if farm else "Farm"
    flock_map = {str(f.id): f for f in data["flocks"]}

    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # ── Colour helpers ────────────────────────────────────────────────────────
    GREEN_FILL = PatternFill("solid", fgColor="16A34A")
    LIGHT_FILL = PatternFill("solid", fgColor="F0FDF4")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(size=9)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header(ws, headers: list[str], widths: list[int]) -> None:
        ws.append(headers)
        for col_i, (cell, w) in enumerate(
            zip(ws[ws.max_row], widths), start=1
        ):
            cell.fill = GREEN_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_i)].width = w

    def write_row(ws, values: list, row_num: int) -> None:
        ws.append(values)
        fill = LIGHT_FILL if row_num % 2 == 0 else PatternFill()
        for cell in ws[ws.max_row]:
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER
            cell.fill = fill

    def freeze_and_filter(ws) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 1: Farm Overview ────────────────────────────────────────────────
    ws = wb.create_sheet("Farm Overview")
    ws.append(["Greena Farm Report"])
    ws["A1"].font = Font(bold=True, size=16, color="14532D")
    ws.append([f"Farm: {farm_name}"])
    ws.append([f"Exported: {data['exported_at'].strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([])
    active = [f for f in data["flocks"] if f.status == "active"]
    ws.append(["Total Flocks (all time)", len(data["flocks"])])
    ws.append(["Active Flocks", len(active)])
    ws.append(["Total Active Birds",
                sum(getattr(f, "current_bird_count", f.initial_bird_count) or 0
                    for f in active)])
    ws.append(["Total Expenses (KES)",
                sum(float(e.amount) for e in data["expenses"])])
    ws.append(["Total Revenue (KES)",
                sum(float(r.amount) for r in data["revenues"])])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: Flocks ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Flocks")
    write_header(ws,
                 ["Flock Name", "Status", "Breed", "Initial Birds",
                  "Current Birds", "Placement Date", "Days Alive",
                  "Expected Cycle Days"],
                 [20, 12, 15, 14, 14, 16, 12, 18])
    for i, f in enumerate(data["flocks"], 1):
        days = None
        if f.placement_date:
            days = (datetime.utcnow().date() - f.placement_date).days
        write_row(ws, [
            f.name, f.status,
            getattr(f, "breed", "") or "",
            f.initial_bird_count,
            getattr(f, "current_bird_count", f.initial_bird_count) or f.initial_bird_count,
            str(f.placement_date) if f.placement_date else "",
            days or "",
            getattr(f, "expected_cycle_days", "") or "",
        ], i)
    freeze_and_filter(ws)

    # ── Sheet 3: Daily Logs ───────────────────────────────────────────────────
    ws = wb.create_sheet("Daily Logs")
    write_header(ws,
                 ["Date", "Flock Name", "Flock ID", "Mortality",
                  "Feed (kg)", "Water (L)", "Morning Count"],
                 [14, 20, 36, 10, 10, 10, 14])
    for i, log in enumerate(data["daily_logs"], 1):
        flock = flock_map.get(str(log.flock_id))
        write_row(ws, [
            str(log.log_date),
            flock.name if flock else "",
            str(log.flock_id),
            log.mortality_count,
            float(log.feed_kg) if log.feed_kg else "",
            float(log.water_litres) if getattr(log, "water_litres", None) else "",
            getattr(log, "morning_count", "") or "",
        ], i)
    freeze_and_filter(ws)

    # ── Sheet 4: Vaccinations ─────────────────────────────────────────────────
    ws = wb.create_sheet("Vaccinations")
    write_header(ws,
                 ["Date Administered", "Flock Name", "Vaccine Name",
                  "Dose", "Next Due Date", "Notes"],
                 [18, 20, 20, 10, 16, 30])
    for i, v in enumerate(data["vaccinations"], 1):
        flock = flock_map.get(str(v.flock_id))
        write_row(ws, [
            str(v.administered_date),
            flock.name if flock else "",
            v.vaccine_name,
            getattr(v, "dose", "") or "",
            str(v.next_due_date) if v.next_due_date else "",
            getattr(v, "notes", "") or "",
        ], i)
    freeze_and_filter(ws)

    # ── Sheet 5: Expenses ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Expenses")
    write_header(ws,
                 ["Date", "Category", "Amount (KES)", "Payment Method",
                  "Flock", "Notes"],
                 [14, 20, 14, 16, 20, 30])
    for i, e in enumerate(data["expenses"], 1):
        flock = flock_map.get(str(e.flock_id)) if getattr(e, "flock_id", None) else None
        write_row(ws, [
            str(e.expense_date),
            getattr(e, "category", "") or "",
            float(e.amount),
            e.payment_method or "",
            flock.name if flock else "",
            e.notes or "",
        ], i)
    freeze_and_filter(ws)

    # ── Sheet 6: Revenue ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Revenue")
    write_header(ws,
                 ["Date", "Type", "Amount (KES)", "Quantity",
                  "Unit Price (KES)", "Buyer", "Flock", "Notes"],
                 [14, 16, 14, 10, 16, 20, 20, 30])
    for i, r in enumerate(data["revenues"], 1):
        flock = flock_map.get(str(r.flock_id)) if getattr(r, "flock_id", None) else None
        write_row(ws, [
            str(r.sale_date),
            r.revenue_type,
            float(r.amount),
            getattr(r, "quantity", "") or "",
            getattr(r, "unit_price", "") or "",
            getattr(r, "buyer", "") or "",
            flock.name if flock else "",
            getattr(r, "notes", "") or "",
        ], i)
    freeze_and_filter(ws)

    # ── Sheet 7: Financial Snapshots ──────────────────────────────────────────
    ws = wb.create_sheet("P&L Snapshots")
    write_header(ws,
                 ["Flock Name", "Total Revenue (KES)", "Total Expenses (KES)",
                  "Gross Profit (KES)", "Gross Margin %", "FCR",
                  "Cost/Bird (KES)", "Profitable"],
                 [20, 18, 20, 18, 15, 10, 16, 12])
    for i, flock in enumerate(data["flocks"], 1):
        snap = data["snapshots"].get(str(flock.id))
        if snap:
            write_row(ws, [
                flock.name,
                float(snap.total_revenue_kes or 0),
                float(snap.total_expenses_kes or 0),
                float(snap.gross_profit_kes or 0),
                float(snap.gross_margin_pct or 0),
                float(snap.fcr_computed) if snap.fcr_computed else "",
                float(snap.cost_per_bird_kes or 0),
                "Yes" if snap.is_profitable else "No",
            ], i)
        else:
            write_row(ws, [flock.name, 0, 0, 0, 0, "", 0, "No"], i)
    freeze_and_filter(ws)

    # Set workbook properties
    wb.properties.title = f"Greena Report — {farm_name}"
    wb.properties.creator = "Greena Agricultural Operating System"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

async def export_csv(db: AsyncSession, farm_id: uuid.UUID) -> bytes:
    """
    Generate a flat CSV of all daily logs.
    Simple, importable into any spreadsheet or analysis tool.
    """
    data = await _load_farm_data(db, farm_id)
    flock_map = {str(f.id): f for f in data["flocks"]}

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    writer.writerow([
        "date", "flock_name", "flock_id", "flock_status",
        "mortality", "feed_kg", "water_litres", "morning_count",
        "farm_name",
    ])

    farm_name = data["farm"].name if data["farm"] else ""

    for log in data["daily_logs"]:
        flock = flock_map.get(str(log.flock_id))
        writer.writerow([
            str(log.log_date),
            flock.name if flock else "",
            str(log.flock_id),
            flock.status if flock else "",
            log.mortality_count,
            float(log.feed_kg) if log.feed_kg else "",
            float(log.water_litres) if getattr(log, "water_litres", None) else "",
            getattr(log, "morning_count", "") or "",
            farm_name,
        ])

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
